from django.shortcuts import render, redirect
from django.urls import reverse
from .models import UserSurvey
from openpyxl import Workbook
import openpyxl
from django.http.response import HttpResponse
import base64
import json

# Home view.
def home(request):
    return render(request, 'core/home.html', {'numUsers': str(len(UserSurvey.objects.all()) + 1)})

# Results.
def results(request):
    users = UserSurvey.objects.all()
    wb = Workbook()
    ws = wb.active
    ws['B2'] = 'REPORTE DE RESULTADOS'
    ws.merge_cells('B2:F2')
    # DATA
    ws['B4'] = 'NOMBRE'
    ws['C4'] = 'ID'
    ws['D4'] = 'CORREO'
    ws['E4'] = 'EDAD'
    ws['F4'] = 'SEXO'
    ws['G4'] = 'NIÑOS'
    ws['H4'] = 'ESTADO'
    ws['I4'] = 'MUNICIPIO'
    ws['J4'] = 'TIPO DE RESIDENCIA'
    ws['K4'] = 'AÑO DE FACULTAD'
    ws['L4'] = 'ESTADO CIVIL'
    ws['M4'] = 'OCUPACIÓN'
    ws['N4'] = 'SEMESTRE'
    ws['O4'] = '¿TRABAJÓ?'
    ws['P4'] = 'PROBLEMAS GRAVES EN EL TRABAJO'
    ws['Q4'] = 'DIFICULTADES EN EL TRABAJO'
    ws['R4'] = 'ACCESO A INTERNET Y DISPOSITIVOS (DOCTOR)'
    ws['S4'] = 'ACCESO A INTERNET Y DISPOSITIVOS (ESTUDIANTE)'
    ws['T4'] = 'DESEMPEÑO ACADÉMICO'
    ws['U4'] = 'TIPO DE ACTIVIDAD'
    ws['V4'] = 'USO DE CUBREBOCAS'
    ws['W4'] = 'TIPO DE CUBREBOCAS'
    # Questions
    ws['Y4'] = 'ME HA COSTADO MUCHO DESCARGAR LA TENSIÓN'
    ws['Z4'] = 'ME DI CUENTA DE QUE TENÍA LA BOCA SECA'
    ws['AA4'] = 'NO PODÍA SENTIR NINGÚN SENTIMIENTO POSITIVO'
    ws['AB4'] = 'SE ME HIZO DIFÍCIL RESPIRAR'
    ws['AC4'] = 'SE ME HIZO DIFÍCIL TOMAR LA INICIATIVA PARA HACER COSAS'
    ws['AD4'] = 'REACCIONÉ EXAGERADAMENTE EN CIERTAS SITUACIONES'
    ws['AE4'] = 'SENTÍ QUE MIS MANOS TEMBLABAN'
    ws['AF4'] = 'HE SENTIDO QUE ESTABA GASTANDO UNA GRAN CANTIDAD DE ENERGÍA'
    ws['AG4'] = 'ESTABA PREOCUPADO POR SITUACIONES EN LAS CUALES PODÍA TENER PÁNICO O EN LAS QUE PODRÍA HACER EL RIDÍCULO'
    ws['AH4'] = 'HE SENTIDO QUE NO HABÍA NADA QUE ME ILUSIONARA'
    ws['AI4'] = 'ME HE SENTIDO INQUIETO'
    ws['AJ4'] = 'SE ME HIZO DIFÍCIL RELAJARME'
    ws['AK4'] = 'ME SENTÍ TRISTE Y DEPRIMIDO'
    ws['AL4'] = 'NO TOLERÉ NADA QUE NO ME PERMITIERA CONTINUAR CON LO QUE ESTABA HACIENDO'
    ws['AM4'] = 'SENTÍ QUE ESTABA AL PUNTO DE PÁNICO'
    ws['AN4'] = 'NO ME PUDE ENTUSIASMAR POR NADA'
    ws['AO4'] = 'SENTÍ QUE VALÍA MUY POCO COMO PERSONA'
    ws['AP4'] = 'HE TENDIDO A SENTIRME ENFADADO CON FACILIDAD'
    ws['AQ4'] = 'SENTÍ LOS LATIDOS DE MI CORAZÓN A PESAR DE NO HABER HECHO NINGÚN ESFUERZO FÍSICO'
    ws['AR4'] = 'TUVE MIEDO SIN RAZÓN'
    ws['AS4'] = 'SENTÍ QUE LA VIDA NO TENÍA NINGÚN SENTIDO'
    ws['AT4'] = 'DURANTE LOS ÚLTIMOS 7 DÍAS, ¿EN CUÁNTOS REALIZÓ ACTIVIDADES FÍSICAS INTENSAS TALES COMO LEVANTAR PESAS, CLAVAR, HACER EJERCICIOS AERÓBICOS O ANDAR RÁPIDO EN BICICLETA?'
    ws['AU4'] = 'HABITUALMENTE, ¿CUÁNTAS HORAS EN TOTAL DEDICÓ A UNA ACTIVIDAD FÍSICA INTENSA EN UNO DE ESOS DÍAS?'
    ws['AV4'] = 'DURANTE LOS ÚLTIMOS 7 DÍAS, ¿EN CUÁNTOS DÍAS HIZO ACTIVIDADES FÍSICAS MODERADAS COMO TRANSPORTAR PESOS LIVIANOS, ANDAR EN BICICLETA A VELOCIDAD REGULAR, JUGAR TENIS? NO INCLUYA CAMINAR.'
    ws['AW4'] = 'HABITUALMENTE, ¿CUÁNTAS HORAS EN TOTAL DEDICÓ A UNA ACTIVIDAD FÍSICA MODERADA EN UNO DE ESOS DÍAS?'
    ws['AX4'] = 'DURANTE LOS ÚLTIMOS 7 DÍAS, ¿EN CUÁNTOS DÍAS CAMINÓ POR LO MENOS 10 MINUTOS SEGUIDOS?"'
    ws['AY4'] = 'HABITUALMENTE, ¿CUÁNTAS HORAS DEDICÓ A CAMINAR EN UNO DE ESOS DÍAS?'
    ws['AZ4'] = 'DURANTE LOS ÚLTIMOS 7 DÍAS, ¿CUÁNTO TIEMPO PASÓ SENTADO DURANTE UN DÍA HÁBIL?'
    ws['BA4'] = '¿HA SALIDO HABITUALMENTE DE CASA (SU LUGAR DE RESIDENCIA ACTUAL) POR CUESTIONES LABORALES?'
    ws['BB4'] = '¿HA DORMIDO MÁS QUE ANTES?'
    ws['BC4'] = '¿HA VISTO LA TV MÁS QUE ANTES?'
    ws['BD4'] = '¿HA PRACTICADO EJERCICIO FÍSICO DE FORMA REGULAR?'
    ws['BE4'] = '¿HA UTILIZADO LAS REDES SOCIALES MÁS QUE ANTES?'
    ws['BF4'] = '¿HAN CAMBIADO MUCHO SUS RUTINAS?'
    ws['BG4'] = '¿HA UTILIZADO INTERNET MÁS QUE ANTES?'
    ws['BH4'] = '¿HA DEDICADO MÁS TIEMPO QUE ANTES A VER PELÍCULAS, LEER, O JUGAR A VIDEOJUEGOS?'
    ws['BI4'] = '¿HA APROVECHADO PARA REALIZAR ACTIVIDADES EN CASA PARA LAS QUE ANTES NO DISPONÍA DE TIEMPO?'
    ws['BJ4'] = '¿HA MANTENIDO SUS CUIDADOS PERSONALES HABITUALES?'
    ws['BK4'] = '¿UTILIZA SIEMPRE O CASI SIEMPRE GUANTES CUANDO SALE DE CASA?'
    ws['BL4'] = '¿MANTIENE SIEMPRE O CASI SIEMPRE LA DISTANCIA DE SEGURIDAD CON OTRAS PERSONAS FUERA DE CASA (AL MENOS 2 METROS)?'
    ws['BM4'] = '¿CREE QUE SE LAVA O DESINFECTA LAS MANOS CON EXCESIVA FRECUENCIA?'
    ws['BN4'] = '¿DESINFECTA HABITUALMENTE LOS OBJETOS Y SUPERFICIES CON GEL DESINFECTANTE, JABÓN, ALCOHOL, ETC.?'
    ws['BO4'] = '¿TOMA PRECAUCIONES HABITUALMENTE AL LLEGAR DEL SUPERMERCADO, LAVANDO LOS ALIMENTOS, DESINFECTANDO EL TELÉFONO CELULAR O LAS LLAVES, ETC.?'
    ws['BP4'] = '¿SE PROTEGE HABITUALMENTE AL TOCAR ZONAS DE POSIBLE CONTAGIO, COMO MANIJAS, ASCENSORES, LECTORES DE TARJETAS DE CRÉDITO, ETC.?'
    ws['BQ4'] = '¿HA TENIDO DISCUSIONES O CONFLICTOS CON SUS FAMILIARES?'
    ws['BR4'] = '¿LE HA PERTURBADO O ALTERADO NO PODER VER A ALGUNOS DE SUS FAMILIARES QUE VEÍA HABITUALMENTE?'
    ws['BS4'] = '¿LE HA PERTURBADO O ALTERADO NO PODER VER A SUS AMIGOS?'
    ws['BT4'] = '¿LE HA PERTURBADO O ALTERADO NO PODER REALIZAR ACTIVIDADES DE OCIO FUERA DE CASA (VIAJAR, SALIR, ETC.)?'
    ws['BU4'] = '¿HA TENIDO PROBLEMAS PARA REALIZAR ACTIVIDADES DE OCIO EN CASA (LEER, ESCRIBIR, VER PELÍCULAS, VIDEOJUEGOS, ETC.)?'
    ws['BV4'] = '¿LE HA PERTURBADO O ALTERADO NO PODER REALIZAR ACTIVIDADES FÍSICAS FUERA DE CASA (DEPORTE, EJERCICIO FÍSICO, IR AL CAMPO, ETC.)?'
    ws['BW4'] = '¿LE HA PERTURBADO O ALTERADO NO PODER SALIR DE CASA SALVO PARA COSAS MUY NECESARIAS?'
    ws['BX4'] = '¿LE HA PERTURBADO O ALTERADO NO PODER DISPONER DE COSAS QUE NECESITABA?'

    cont = 5

    for user in users:
        ws.cell(row=cont,column=2).value = user.name
        ws.cell(row=cont,column=3).value = user.surveyId
        ws.cell(row=cont,column=4).value = user.email
        ws.cell(row=cont,column=5).value = user.age
        ws.cell(row=cont,column=6).value = user.sex
        ws.cell(row=cont,column=7).value = user.kids
        ws.cell(row=cont,column=8).value = user.state
        ws.cell(row=cont,column=9).value = user.municipality
        ws.cell(row=cont,column=10).value = user.typeOfResidence
        ws.cell(row=cont,column=11).value = user.facultyYear
        ws.cell(row=cont,column=12).value = user.civilStatus
        ws.cell(row=cont,column=13).value = user.occupation
        ws.cell(row=cont,column=14).value = user.semester
        ws.cell(row=cont,column=15).value = user.worked
        ws.cell(row=cont,column=16).value = user.seriousWorkProblems
        ws.cell(row=cont,column=17).value = user.workDifficulties
        ws.cell(row=cont,column=18).value = user.accessToInternetAndDevices_Doctor
        ws.cell(row=cont,column=19).value = user.accessToInternetAndDevices_Student
        ws.cell(row=cont,column=20).value = user.academicPerformance
        ws.cell(row=cont,column=21).value = user.difficultiesToStudy
        ws.cell(row=cont,column=22).value = user.typeOfActivity
        ws.cell(row=cont,column=23).value = user.useOfMask
        ws.cell(row=cont,column=24).value = user.typeOfMask
        ws.cell(row=cont,column=26).value = user.pt1_1
        ws.cell(row=cont,column=27).value = user.pt1_2
        ws.cell(row=cont,column=28).value = user.pt1_3
        ws.cell(row=cont,column=29).value = user.pt1_4
        ws.cell(row=cont,column=30).value = user.pt1_5
        ws.cell(row=cont,column=31).value = user.pt1_6
        ws.cell(row=cont,column=32).value = user.pt1_7
        ws.cell(row=cont,column=33).value = user.pt1_8
        ws.cell(row=cont,column=34).value = user.pt1_9
        ws.cell(row=cont,column=35).value = user.pt1_10
        ws.cell(row=cont,column=36).value = user.pt1_11
        ws.cell(row=cont,column=37).value = user.pt1_12
        ws.cell(row=cont,column=38).value = user.pt1_13
        ws.cell(row=cont,column=39).value = user.pt1_14
        ws.cell(row=cont,column=40).value = user.pt1_15
        ws.cell(row=cont,column=41).value = user.pt1_16
        ws.cell(row=cont,column=42).value = user.pt1_17
        ws.cell(row=cont,column=43).value = user.pt1_18
        ws.cell(row=cont,column=44).value = user.pt1_19
        ws.cell(row=cont,column=45).value = user.pt1_20
        ws.cell(row=cont,column=46).value = user.pt1_21
        ws.cell(row=cont,column=47).value = user.pt2_Intense_1
        ws.cell(row=cont,column=48).value = user.pt2_Intense_2
        ws.cell(row=cont,column=49).value = user.pt2_Moderated_1
        ws.cell(row=cont,column=50).value = user.pt2_Moderated_2
        ws.cell(row=cont,column=51).value = user.pt2_Hike_1
        ws.cell(row=cont,column=52).value = user.pt2_Hike_2
        ws.cell(row=cont,column=53).value = user.pt2_Seated_1
        ws.cell(row=cont,column=54).value = user.pt3_1
        ws.cell(row=cont,column=55).value = user.pt3_2
        ws.cell(row=cont,column=56).value = user.pt3_3
        ws.cell(row=cont,column=57).value = user.pt3_4
        ws.cell(row=cont,column=58).value = user.pt3_5
        ws.cell(row=cont,column=59).value = user.pt3_6
        ws.cell(row=cont,column=60).value = user.pt3_7
        ws.cell(row=cont,column=61).value = user.pt3_8
        ws.cell(row=cont,column=62).value = user.pt3_9
        ws.cell(row=cont,column=63).value = user.pt3_10
        ws.cell(row=cont,column=64).value = user.pt4_1
        ws.cell(row=cont,column=65).value = user.pt4_2
        ws.cell(row=cont,column=66).value = user.pt4_3
        ws.cell(row=cont,column=67).value = user.pt4_4
        ws.cell(row=cont,column=68).value = user.pt4_5
        ws.cell(row=cont,column=69).value = user.pt4_6
        ws.cell(row=cont,column=70).value = user.pt5_1
        ws.cell(row=cont,column=71).value = user.pt5_2
        ws.cell(row=cont,column=72).value = user.pt5_3
        ws.cell(row=cont,column=73).value = user.pt5_4
        ws.cell(row=cont,column=74).value = user.pt5_5
        ws.cell(row=cont,column=75).value = user.pt5_6
        cont+=1

    filename = 'Reporte_De_Resultados.xlsx'
    response = HttpResponse(content_type = 'application/ms-excel')
    content = 'attachment; filename = {0}'.format(filename)
    response['Content-Disposition'] = content
    wb.save(response)

    return response

# Format the name with the first letter capitalized in each word.
def formatName(name):
    # Separate the name.
    initialsName = name.split(' ')
    # Capitalize de first letter of each word.
    formated = []
    for name in initialsName:
        formated.append(name.lower().capitalize())
    # Order the name
    if len(formated) > 3:
        formatedName = formated[0] + ' ' + formated[1] + ' ' + formated[2] + ' ' + formated[3]
    else:
        formatedName = formated[0] + ' ' + formated[1] + ' ' + formated[2]
    return formatedName

# Generate the ID for the new user.
def getID(name, facultyYear):
    # Get the folio by knowing the number of users.
    folio = str(len(UserSurvey.objects.all()) + 1)
    # Format the name and get its initials.
    initials = formatName(name).split(' ')
    # Get the abreviation of the name.
    if len(initials) > 3:
        abrev = initials[2][0] + initials[3][0] + initials[0][0] + initials[1][0]
    else:
        abrev = initials[1][0] + initials[2][0] + initials[0][0] + '0'
    # Generate the id.
    surveyId = facultyYear + abrev + '_' + folio
    return surveyId

# Decode the image in base64 and saves it as the ID value.
def decodeImage(name, facultyYear, image_base64):
    # Decode image in base64 format.
    imgdata = base64.b64decode(image_base64)
    # Format the name of the new image.
    filename = 'usersSignature/' + getID(name, facultyYear) + '.jpg'
    # Open file and save it in 'media' directory.
    with open('media/'+filename, 'wb') as f:
        f.write(imgdata)
    return filename

# Check the format of the image.
def imageFormat(base64):
    if 'jpeg' in base64:
        return base64.replace('data:image/jpeg;base64,', '')
    if 'jpg' in base64:
        return base64.replace('data:image/jpg;base64,', '')
    if 'png' in base64:
        return base64.replace('data:image/png;base64,', '')

# Get all the data from the request.
def getQuestions(request):
    # Questions part 1.
    pt1_1 = request.POST['pt1_1']
    pt1_2 = request.POST['pt1_2']
    pt1_3 = request.POST['pt1_3']
    pt1_4 = request.POST['pt1_4']
    pt1_5 = request.POST['pt1_5']
    pt1_6 = request.POST['pt1_6']
    pt1_7 = request.POST['pt1_7']
    pt1_8 = request.POST['pt1_8']
    pt1_9 = request.POST['pt1_9']
    pt1_10 = request.POST['pt1_10']
    pt1_11 = request.POST['pt1_11']
    pt1_12 = request.POST['pt1_12']
    pt1_13 = request.POST['pt1_13']
    pt1_14 = request.POST['pt1_14']
    pt1_15 = request.POST['pt1_15']
    pt1_16 = request.POST['pt1_16']
    pt1_17 = request.POST['pt1_17']
    pt1_18 = request.POST['pt1_18']
    pt1_19 = request.POST['pt1_19']
    pt1_20 = request.POST['pt1_20']
    pt1_21 = request.POST['pt1_21']
    # Questions part 2.
    pt2_Intense_1 = request.POST['pt2_Intense_1']
    pt2_Intense_2 = request.POST['pt2_Intense_2']
    pt2_Moderated_1 = request.POST['pt2_Moderated_1']
    pt2_Moderated_2 = request.POST['pt2_Moderated_2']
    pt2_Hike_1 = request.POST['pt2_Hike_1']
    pt2_Hike_2 = request.POST['pt2_Hike_2']
    pt2_Seated_1 = request.POST['pt2_Seated_1']
    # Questions part 3.
    pt3_1 = request.POST['pt3_1']
    pt3_2 = request.POST['pt3_2']
    pt3_3 = request.POST['pt3_3']
    pt3_4 = request.POST['pt3_4']
    pt3_5 = request.POST['pt3_5']
    pt3_6 = request.POST['pt3_6']
    pt3_7 = request.POST['pt3_7']
    pt3_8 = request.POST['pt3_8']
    pt3_9 = request.POST['pt3_9']
    pt3_10 = request.POST['pt3_10']
    # Questions part 4.
    pt4_1 = request.POST['pt4_1']
    pt4_2 = request.POST['pt4_2']
    pt4_3 = request.POST['pt4_3']
    pt4_4 = request.POST['pt4_4']
    pt4_5 = request.POST['pt4_5']
    pt4_6 = request.POST['pt4_6']
    # Questions part 5.
    pt5_1 = request.POST['pt5_1']
    pt5_2 = request.POST['pt5_2']
    pt5_3 = request.POST['pt5_3']
    pt5_4 = request.POST['pt5_4']
    pt5_5 = request.POST['pt5_5']
    pt5_6 = request.POST['pt5_6']
    pt5_5 = request.POST['pt5_5']
    pt5_6 = request.POST['pt5_6']
    pt5_7 = request.POST['pt5_7']
    pt5_8 = request.POST['pt5_8']

    return (pt1_1, pt1_2, pt1_3, pt1_4, pt1_5, pt1_6, pt1_7, pt1_8, pt1_9, pt1_10, pt1_11, pt1_12, pt1_13, pt1_14, pt1_15, pt1_16, pt1_17, pt1_18, pt1_19, pt1_20, pt1_21, pt2_Intense_1, pt2_Intense_2, pt2_Moderated_1, pt2_Moderated_2, pt2_Hike_1, pt2_Hike_2, pt2_Seated_1, pt3_1, pt3_2, pt3_3, pt3_4, pt3_5, pt3_6, pt3_7, pt3_8, pt3_9, pt3_10, pt4_1, pt4_2, pt4_3, pt4_4, pt4_5, pt4_6, pt5_1, pt5_2, pt5_3, pt5_4, pt5_5, pt5_6, pt5_7, pt5_8)

# Get all the data from the request.
def getGeneralData(request):
    facultyYear = request.POST['facultyYear']

    # Name.
    try:
        name = formatName(request.POST['name'])
    except:
        return redirect(reverse('home') + '?name')
    
    # Image.
    image_base64 = request.POST['image']
    try:
        image = decodeImage(name, facultyYear, imageFormat(image_base64))
    except:
        return redirect(reverse('home') + '?image')

    # ID.
    surveyId = getID(name, facultyYear)

    # General data.
    age = request.POST['age']
    sex = request.POST['sex']
    kids = request.POST['kids']
    state = request.POST['state']
    civilStatus = request.POST['civilStatus']
    municipality = request.POST['municipality'].lower().capitalize()
    typeOfResidence = request.POST['typeOfResidence']
    email = request.POST['email']
    occupation = request.POST['occupation']
    typeOfActivity = request.POST['typeOfActivity']
    useOfMask = request.POST['useOfMask']
    typeOfMask = request.POST['typeOfMask']
    semester = request.POST['semester']
    worked = request.POST['worked']
    seriousWorkProblems = request.POST['seriousWorkProblems']
    workDifficulties = request.POST['workDifficulties']
    accessToInternetAndDevices_Doctor = request.POST['accessToInternetAndDevices_Doctor']
    accessToInternetAndDevices_Student = request.POST['accessToInternetAndDevices_Student']
    academicPerformance = request.POST['academicPerformance']
    difficultiesToStudy = request.POST['difficultiesToStudy']

    return (name, surveyId, email, age, sex, kids, state, municipality, typeOfResidence, facultyYear, civilStatus, occupation, semester, worked, seriousWorkProblems, workDifficulties, accessToInternetAndDevices_Doctor, accessToInternetAndDevices_Student, academicPerformance, difficultiesToStudy, typeOfActivity, useOfMask, typeOfMask, image)

# Change value 'undefined' for 'No aplica'.
def formatUndefinedField(dataList):
    for item in dataList:
        if item == 'undefined':
            dataList[dataList.index(item)] = 'No aplica'
    return tuple(dataList)

def saveForm(request):
    if request.method == 'POST':

        # Get all the data.
        data = formatUndefinedField(list(getGeneralData(request)))
        # Get all the questions.
        questions = formatUndefinedField(list(getQuestions(request)))
        
        # If email doen't exists, create new user.
        if not UserSurvey.objects.filter(email=data[2]).exists():
            try:
                # Create a new user.
                user = UserSurvey(
                    name=data[0],
                    surveyId=data[1],
                    email=data[2],
                    age=data[3],
                    sex=data[4],
                    kids=data[5],
                    state=data[6],
                    municipality=data[7],
                    typeOfResidence=data[8],
                    facultyYear=data[9],
                    civilStatus=data[10],
                    occupation=data[11],
                    semester=data[12],
                    worked=data[13],
                    seriousWorkProblems=data[14],
                    workDifficulties=data[15],
                    accessToInternetAndDevices_Doctor=data[16],
                    accessToInternetAndDevices_Student=data[17],
                    academicPerformance=data[18],
                    difficultiesToStudy=data[19],
                    typeOfActivity=data[20],
                    useOfMask=data[21],
                    typeOfMask=data[22],
                    image=data[23],
                    pt1_1=questions[0],
                    pt1_2=questions[1],
                    pt1_3=questions[2],
                    pt1_4=questions[3],
                    pt1_5=questions[4],
                    pt1_6=questions[5],
                    pt1_7=questions[6],
                    pt1_8=questions[7],
                    pt1_9=questions[8],
                    pt1_10=questions[9],
                    pt1_11=questions[10],
                    pt1_12=questions[11],
                    pt1_13=questions[12],
                    pt1_14=questions[13],
                    pt1_15=questions[14],
                    pt1_16=questions[15],
                    pt1_17=questions[16],
                    pt1_18=questions[17],
                    pt1_19=questions[18],
                    pt1_20=questions[19],
                    pt1_21=questions[20],
                    pt2_Intense_1=questions[21],
                    pt2_Intense_2=questions[22],
                    pt2_Moderated_1=questions[23],
                    pt2_Moderated_2=questions[24],
                    pt2_Hike_1=questions[25],
                    pt2_Hike_2=questions[26],
                    pt2_Seated_1=questions[27],
                    pt3_1=questions[28],
                    pt3_2=questions[29],
                    pt3_3=questions[30],
                    pt3_4=questions[31],
                    pt3_5=questions[32],
                    pt3_6=questions[33],
                    pt3_7=questions[34],
                    pt3_8=questions[35],
                    pt3_9=questions[36],
                    pt3_10=questions[37],
                    pt4_1=questions[38],
                    pt4_2=questions[39],
                    pt4_3=questions[40],
                    pt4_4=questions[41],
                    pt4_5=questions[42],
                    pt4_6=questions[43],
                    pt5_1=questions[44],
                    pt5_2=questions[45],
                    pt5_3=questions[46],
                    pt5_4=questions[47],
                    pt5_5=questions[48],
                    pt5_6=questions[49],
                    pt5_7=questions[50],
                    pt5_8=questions[51]
                )
                # Save the user in DB.
                user.save()
                return redirect(reverse('home') + '?ok')
            except:
                return redirect(reverse('home') + '?fail')
        else:
            return redirect(reverse('home') + '?alreadyUsed')
        
    return render(request, 'core/home.html')